import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
# wb.create_sheet("Sheet2", "0")

cell = sheet["a1"]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

cell = sheet.cell(row=1, column=1)

# print(sheet.max_row)
# print(sheet.max_column)
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)


# cells = sheet["a"]
# print(cells)

# cells_a_c = sheet["a:c"]
# print(cells_a_c)

cells_a1_c3 = sheet["a1:c3"]
print(cells_a1_c3)

# print(sheet[1])  # all cels of first row

sheet.append([1, 2, 3])
# sheet.insert_rows()
# sheet.delete_cols()

wb.save("transactions2.xlsx")
